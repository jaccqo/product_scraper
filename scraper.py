import undetected_chromedriver as uc
import time
from termcolor import colored
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import random
from selenium.common.exceptions import TimeoutException,ElementClickInterceptedException,ElementNotInteractableException,StaleElementReferenceException,NoSuchElementException
import bs4
import requests
import colorama
import datetime as dt
# import openpyxl module
import openpyxl
import os
import csv


title=['Handle', 'Title', 'Body (HTML)', 'Vendor', 'Product Category', 'Type', 'Tags', 'Published', 'Option1 Name', 'Option1 Value', 'Option2 Name', 'Option2 Value', 'Option3 Name', 'Option3 Value', 'Variant SKU', 
       'Variant Grams', 'Variant Inventory Tracker', 'Variant Inventory Qty', 'Variant Inventory Policy', 'Variant Fulfillment Service', 'Variant Price', 'Variant Compare At Price', 'Variant Requires Shipping',
         'Variant Taxable', 'Variant Barcode', 'Image Src', 'Image Position', 'Image Alt Text', 'Gift Card', 'SEO Title', 'SEO Description', 'Google Shopping / Google Product Category', 'Google Shopping / Gender',
           'Google Shopping / Age Group', 'Google Shopping / MPN', 'Google Shopping / AdWords Grouping', 'Google Shopping / AdWords Labels', 'Google Shopping / Condition', 'Google Shopping / Custom Product', 
           'Google Shopping / Custom Label 0', 'Google Shopping / Custom Label 1', 'Google Shopping / Custom Label 2', 'Google Shopping / Custom Label 3', 'Google Shopping / Custom Label 4', 'Variant Image', 
           'Variant Weight Unit', 'Variant Tax Code', 'Cost per item', 'Price / International', 'Compare At Price / International', 'Status']



colorama.init()

class Excel:
    def __init__(self,img_url,product_title,price,part_number,description,status):
        file_name="new_product.xlsx"
        skip_product=False
        if os.path.exists(file_name):
            wb=openpyxl.load_workbook(file_name)
            sheet=wb.active
            # search for products with same title
            search=[sheet[f"B{k}"].value for k in range(1,sheet.max_row+1)]
           
            if product_title in search:
                skip_product=True
            del search

        else:
           
            wb = openpyxl.Workbook()

            sheet = wb.active
            sheet.append(("IMAGE URL","PRODUCT TITLE","PRICE","PART NUMBER","DESCRIPTION","STATUS"))

        if not skip_product:
            sheet.append((img_url,product_title,price,part_number,description,status))

            wb.save(file_name)
        else:
            print(colored('[  ]product already exists in excel',"green"))


class convert_to_csv:
    def __init__(self):
        pass 
    def convert_file(self):

        file_name="new_product.xlsx"
        wb=openpyxl.load_workbook(file_name)
        sheet=wb.active
        # search for products with same title

        # image urls
        main_rows=[]
        
        
        csv_file=open("new_product.csv","w", encoding='utf-8')
        csv_writer=csv.writer(csv_file,delimiter=",",lineterminator='\r')

        csv_writer.writerow(title)


       
        column_a="Image Src"
        column_a_rows=[]
        for k in range(2,sheet.max_row+1):
            row_vals=['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'active']

            print(colored(f"creating image sources {round(k/sheet.max_row+1,4*100)}%","green"))
           

            val_position=title.index(column_a)

            row_vals[val_position]= sheet[f"A{k}"].value

            column_a_rows.append(row_vals)

            #csv_writer.writerow(row_vals)



        # product title
        column_b="Title"
        column_b_rows=[]
        column_zero="Handle"
        for k in range(2,sheet.max_row+1):
            print(colored(f"creating titles {round(k/sheet.max_row+1*100,4)}%","green"))
        

            row_vals=['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'active']

            val_position=title.index(column_b)

            row_vals[val_position]= sheet[f"B{k}"].value

            
            val_position=title.index(column_zero)

            row_vals[val_position]= sheet[f"B{k}"].value

            column_b_rows.append(row_vals)


           

           # csv_writer.writerow(row_vals)
        
        # price
        column_c="Variant Price"
        colum_c_rows=[]
        for k in range(2,sheet.max_row+1):
            print(colored(f"creating prices {round(k/sheet.max_row+1*100,4)}%","green"))
           
            row_vals=['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'active']

            val_position=title.index(column_c)

            row_vals[val_position]= sheet[f"C{k}"].value

            colum_c_rows.append(row_vals)


            #csv_writer.writerow(row_vals)
        
        # part number
        column_d="Variant SKU"
        column_d_rows=[]
        for k in range(2,sheet.max_row+1):
            print(colored(f"creating sku numbers {round(k/sheet.max_row+1*100,4)}%","green"))

            row_vals=['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'active']

            val_position=title.index(column_d)

            row_vals[val_position]= sheet[f"D{k}"].value

            column_d_rows.append(row_vals)


         #description
        column_e="SEO Description"
        column_e_rows=[]
        options = uc.ChromeOptions()
        # options.headless=True
        # options.add_argument('--headless')

        self.driver2 = uc.Chrome(options=options) 

        self.driver2.get("https://translate.google.co.il/?sl=auto&tl=iw&op=translate")
        max_rows=sheet.max_row+1
        start_time=time.time()
        for k in range(2,max_rows):
           
            desc=(sheet[f"E{k}"].value)

            hebrew_desc=self.hebrew_translator(desc,max_rows,start_time,k)

            row_vals=['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'active']

            val_position=title.index(column_e)

            row_vals[val_position]= hebrew_desc
            column_e_rows.append(row_vals)


        self.driver2.close()

        for rows in range(len(column_a_rows)):
            row_vals=['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'deny', 'manual', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'active']

            for column in range(len(column_a_rows[rows])):
                if column_a_rows[rows][column]:
                    row_vals[column]=column_a_rows[rows][column]
                if column_b_rows[rows][column]:
                    row_vals[column]=column_b_rows[rows][column].strip().replace("\n"," ")
                if column_d_rows[rows][column]:
                    row_vals[column]=column_d_rows[rows][column]
                if column_d_rows[rows][column]:
                    row_vals[column]=column_d_rows[rows][column]
                if column_e_rows[rows][column]:
                    row_vals[column]=column_e_rows[rows][column].strip().replace("\n"," ")
            " write each row"
           
            csv_writer.writerow(row_vals)

        csv_file.close()

        print("program finished with zero errors")

    def hebrew_translator(self,desc,enum_trans,start_time,current_iter):
    
        print("")
        k=current_iter+1
        print(colored(f"{time.ctime()} translation progress {round(k/enum_trans*100,4)}% {self.calcProcessTime(start_time,k,enum_trans)}","red"))
        print(f"{desc} ")

        selected_elements=WebDriverWait(self.driver2, 2000).until(
        EC.presence_of_element_located((By.CSS_SELECTOR,
                                        'textarea[aria-label="טקסט מקור"]')))
        
        selected_elements.send_keys(desc)

        time.sleep(3)

        translated=WebDriverWait(self.driver2, 2000).until(
        EC.presence_of_element_located((By.CSS_SELECTOR,
                                        'div[class="lRu31"]')))
        

        print(colored(f"[ ] {translated.text}","green"))

    


        selected_elements.clear()
        time.sleep(1)

        return translated.text


    
    def calcProcessTime(self,starttime, cur_iter, max_iter):

        telapsed = time.time() - starttime
        testimated = (telapsed/cur_iter)*(max_iter)

        finishtime = starttime + testimated
        finishtime = f'eta {dt.datetime.fromtimestamp(finishtime).strftime("%H:%M:%S")}'  # in time

        lefttime = testimated-telapsed  # in seconds

        time_left=f"remaining/time {dt.timedelta(seconds=lefttime)}"

        return (time_left, finishtime)
        



class shopify_scraper:

    def __init__(self):
        options = uc.ChromeOptions()
        # setting profile
        # options.user_data_dir = "c:\\temp\\profile"

        # use specific (older) version
        self.driver = uc.Chrome(options=options) 

    def start(self):

        url="https://www.quadratec.com/categories/jeep-overland-camping-gear"

        print(colored(f"Getting {url}","green"))

        self.driver.get(url)

        selected_elements=WebDriverWait(self.driver, 20).until(
        EC.presence_of_all_elements_located((By.CSS_SELECTOR,
                                        "li[class='col-xs-4 col-sm-3 col-md-2 text-center']")))
        
        
        print(selected_elements,end="\n")
        elems=0
        # loop through each camping item in website 
       
        while elems<len(selected_elements):
            k=elems+1
            try:
                selected_elements[elems].click()

                time.sleep(0.5)

            
                print(colored(f"collecting data  {round(k/len(selected_elements)*100,4)}%","green"))
                elems+=1

                # get each item inside the collection

                # collection len
                current_url=""
                self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
                try:

                    pages_count=WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR,
                                                "ul[class='pagination']")))
                    
                except (TimeoutException, ElementClickInterceptedException, ElementNotInteractableException,
                        StaleElementReferenceException, NoSuchElementException) as e:
                    pages_count=1
               
                    
            except (TimeoutException, ElementClickInterceptedException, ElementNotInteractableException,
                        StaleElementReferenceException, NoSuchElementException) as e:
                selected_elements[elems].click()

                time.sleep(0.5)

            
                print(colored(f"collecting data  {round(k/len(selected_elements)*100,4)}%","green"))
                elems+=1

                # get each item inside the collection

                # collection len
                current_url=""
                self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")

                try:

                    pages_count=WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR,
                                                "ul[class='pagination']")))
                    
                except (TimeoutException, ElementClickInterceptedException, ElementNotInteractableException,
                        StaleElementReferenceException, NoSuchElementException) as e:
                    pages_count=1
                    
                
            
            if pages_count==1:
                pages_len=1
            else:
                pages_count=bs4.BeautifulSoup(pages_count.get_attribute("innerHTML"),"lxml")
                pages_len=pages_count.find_all("li")
                pages_len=len(pages_len)-2

            print(colored(f" total pages found {pages_len}","red"))
            page=1
            while page<=pages_len:
                print(colored(f" pages progress {round(page/pages_len*100,4)}","red"))
                try:
                    if page==1:
                        collection_len=WebDriverWait(self.driver, 20).until(
                        EC.presence_of_all_elements_located((By.CSS_SELECTOR,
                                                    "div[class='row row-autowrap']")))[1]
                        
                        self.driver.execute_script("arguments[0].scrollIntoView();", collection_len)


                        soup=bs4.BeautifulSoup(collection_len.get_attribute("innerHTML"),"lxml")

                        collection_articles = soup.find_all("article")

                        current_collection_len=len(collection_articles)

                        print(colored(f"This collection len {current_collection_len}","cyan"))

                    else:
                        collection_len=WebDriverWait(self.driver, 20).until(
                        EC.presence_of_all_elements_located((By.CSS_SELECTOR,
                                                    "div[class='row row-autowrap']")))[0]
                        
                        self.driver.execute_script("arguments[0].scrollIntoView();", collection_len)


                        soup=bs4.BeautifulSoup(collection_len.get_attribute("innerHTML"),"lxml")

                        collection_articles = soup.find_all("article")

                        current_collection_len=len(collection_articles)

                        print(colored(f"This collection len {current_collection_len}","cyan"))

                except (TimeoutException, ElementClickInterceptedException, ElementNotInteractableException,
                        StaleElementReferenceException, NoSuchElementException) as e:
                    
                    if page==1:
                        collection_len=WebDriverWait(self.driver, 20).until(
                        EC.presence_of_all_elements_located((By.CSS_SELECTOR,
                                                    "div[class='row row-autowrap']")))[1]
                        
                        self.driver.execute_script("arguments[0].scrollIntoView();", collection_len)


                        soup=bs4.BeautifulSoup(collection_len.get_attribute("innerHTML"),"lxml")

                        collection_articles = soup.find_all("article")

                        current_collection_len=len(collection_articles)

                        print(colored(f"This collection len {current_collection_len}","cyan"))

                    else:
                        collection_len=WebDriverWait(self.driver, 20).until(
                        EC.presence_of_all_elements_located((By.CSS_SELECTOR,
                                                    "div[class='row row-autowrap']")))[0]
                        
                        self.driver.execute_script("arguments[0].scrollIntoView();", collection_len)


                        soup=bs4.BeautifulSoup(collection_len.get_attribute("innerHTML"),"lxml")

                        collection_articles = soup.find_all("article")

                        current_collection_len=len(collection_articles)

                        print(colored(f"This collection len {current_collection_len}","cyan"))
                
                item=1
            
                #get each item in columns
                while item<current_collection_len:
                    try:
                        try:
                            collection_item=WebDriverWait(self.driver,10).until(
                            EC.presence_of_element_located((By.XPATH,f"/html/body/div[4]/div/div/section/div[2]/section[3]/div/div/div/div[3]/div/div/div[1]/article[{item}]")))
                        except (TimeoutException, ElementClickInterceptedException, ElementNotInteractableException,
                        StaleElementReferenceException, NoSuchElementException) as e:
                            print("retrying..")
                            if page==1:
                                collection_item=WebDriverWait(self.driver,5).until(
                                EC.presence_of_element_located((By.XPATH,f"/html/body/div[4]/div/div/section/div[2]/section[3]/div/div/div/div[2]/div/div/div[1]/article[{item}]")))
                            else:
                                collection_item=WebDriverWait(self.driver,5).until(EC.presence_of_element_located((By.XPATH, f"/html/body/div[4]/div/div/section/div[2]/section[3]/div/div/div/div/div/div/div[1]/article[{item}]")))
                 
                                                                    
                                                                   
                    
                    except (TimeoutException, ElementClickInterceptedException, ElementNotInteractableException,
                        StaleElementReferenceException, NoSuchElementException) as e:
                        try:
                            collection_item=WebDriverWait(self.driver,10).until(
                            EC.presence_of_element_located((By.XPATH,f"/html/body/div[4]/div/div/section/div[2]/section[3]/div/div/div/div[3]/div/div/div[1]/article[{item}]")))
                        except (TimeoutException, ElementClickInterceptedException, ElementNotInteractableException,
                        StaleElementReferenceException, NoSuchElementException) as e:
                            print("retrying..")
                            if page==1:
                                collection_item=WebDriverWait(self.driver,5).until(
                                EC.presence_of_element_located((By.XPATH,f"/html/body/div[4]/div/div/section/div[2]/section[3]/div/div/div/div[2]/div/div/div[1]/article[{item}]")))
                            else:
                                collection_item=WebDriverWait(self.driver,5).until(EC.presence_of_element_located((By.XPATH, f"/html/body/div[4]/div/div/section/div[2]/section[3]/div/div/div/div/div/div/div[1]/article[{item}]")))
                 
                                           
                    self.driver.execute_script("arguments[0].scrollIntoView();", collection_item)

                    item_attributes=bs4.BeautifulSoup(collection_item.get_attribute("innerHTML"),"lxml")
                    

                    image_link=item_attributes.find('img')["src"]            
                    title=item_attributes.find('div',{"class":"title"})
                    price=item_attributes.find('span',{"class":"product-price"})
                    suggested_status=item_attributes.find('span',{"class":"suggested-price"})
                    stock_status=item_attributes.find('div',{"class":"serp-stock-status"})
                    part_number_url=item_attributes.find('a')['href']

                    part_number_url=f"https://www.quadratec.com{part_number_url}"

                    part_num=requests.get(part_number_url).text

                    parsed_num=bs4.BeautifulSoup(part_num,"lxml")

                    img_url=parsed_num.find("div",{"class":"slide__content"})
                    
                    if img_url:
                        if ".jpg" not in img_url:
                            
                            img_url=parsed_num.find("div",{"class":"slick__slide"})
                        
                        try:
                            img_url=img_url.find("a")["href"]
                        except Exception as e:
                            img_url="could not get image"

                    else:
                        img_url="could not get image"


                    product_description=parsed_num.find('div',{"itemprop":"description"})
                    
                    parsed_num=parsed_num.find('span',{"class":"num-value"})

                    product_title=title.get_text()
                    product_price=price.get_text()
                    if parsed_num:
                        product_num=parsed_num.get_text()
                    else:
                        product_num="none"


                   
                    print(colored(product_title,"yellow"))
                    print(product_price)
                    if suggested_status:
                        print(suggested_status.get_text())
                    print(stock_status.get_text())
                    print(product_num)

                    print(img_url)
            
                  
                    if product_description:
                        product_desc=product_description.get_text()

                        
                    else:
                        product_desc="no description found"

                    print(colored(product_desc,"cyan"))

                    print("\n\n")
                    image_link=img_url

                    Excel(image_link,product_title,product_price,product_num,product_desc,stock_status.get_text())
                
                    time.sleep(2)
                    item+=1

                if page==1:

                    current_url=self.driver.current_url

                print(page,"||",pages_len)

                if page==pages_len:
                    print("reached end of resuslts")
                    break
                else:
                    self.driver.get(f"{current_url}?page={page}")

                    page+=1

    
            print(colored("going back to main page","cyan"))
            self.driver.get(url)

            selected_elements=WebDriverWait(self.driver, 20).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR,
                                        "li[class='col-xs-4 col-sm-3 col-md-2 text-center']")))
                
        
        

        self.driver.close()

if __name__=="__main__":
    
    # scrape=shopify_scraper()
    # scrape.start()
    
    convert_to_csv().convert_file()


